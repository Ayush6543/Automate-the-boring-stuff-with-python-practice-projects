# ! python3

"""
MultiplicationTableMaker.py - A program that takes N number from command line
and creates a NxN multiplication table in an Excel spreadsheet.
"""


def multiplication_table():
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    sheet = wb.get_sheet_by_name('Sheet')

    cell_obj = Font(bold=True)

    for i in range(1, 7):
        sheet.cell(row=i + 1, column=1).value = i
        sheet.cell(row=i + 1, column=1).font = cell_obj
    for i in range(1, 7):
        sheet.cell(row=1, column=i + 1).value = i
        sheet.cell(row=1, column=i + 1).font = cell_obj
    for row in range(1, 7):
        for col in range(1, 7):
            data = sheet.cell(row=row + 1, column=col + 1).value = row * col

    wb.save('multiplication_table_maker.xlsx')


if __name__ == '__main__':
    multiplication_table()
