# ! python3
# clean_up_data.py - Clean up phone numbers from spreadsheets using regex.

import openpyxl, re

mo = re.compile(r'(\+91?)([7]{1}[0-9]{9})')

wb = openpyxl.load_workbook('clipboardData.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

for row in range(1, sheet.max_row + 1):
    data = sheet.cell(row=row, column=3).value
    # print(data)
    search = mo.search(data)
    # print(search)
    area_code = search.group(1)
    phone_num = search.group(2)

    if search.string == data:
        sheet.cell(row=row, column=3).value = phone_num

wb.save('clipboard_data_copy.xlsx')
