# ! python3
"""
text_files2spread.py - A program to write text files to spreadsheet.
"""

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

file_obj = open('first_row.txt')
file_obj1 = open('second_row.txt')
first_col = file_obj.readlines()
second_col = file_obj1.readlines()

col = 1

for i in range(1, len(first_col) - 1):
    row = 1
    for x in first_col:
        remove_lines = x.strip('\n')
        data = sheet.cell(row=row, column=col).value = remove_lines

        row += 1
    col += 1


for i in range(1, len(second_col) - 1):
    row1 = 1
    for x in second_col:
        lines_remove = x.strip('\n')
        data2 = sheet.cell(row=row1, column=2).value = lines_remove

        row1 += 1
    col += 1

wb.save('text_files2spread.xlsx')
