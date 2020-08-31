# ! python3
"""
spead2txt_file.py - A program to write spreadsheet file to text files.
"""

import openpyxl

wb = openpyxl.load_workbook('text_files2spread.xlsx')
sheet = wb.active

f_col = open('first_col.txt', 'w')
s_col = open('second_col.txt', 'w')

number_of_files = int(input("No. of files being used: "))
col = 1
while col != number_of_files:
    for row in range(1, sheet.max_row + 1):
        data = sheet.cell(row=row, column=col).value
        f_col.write(data + '\n')
    col += 1
    for row in range(1, sheet.max_row + 1):
        data1 = sheet.cell(row=row, column=col).value
        s_col.write(data1 + '\n')
