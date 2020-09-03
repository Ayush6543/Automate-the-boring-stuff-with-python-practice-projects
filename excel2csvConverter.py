# ! python3

"""
excel-to-CSV Converter`.py - A program that converts several excel files to CSV files.
"""

import openpyxl, csv, os


def excel_2_csv_converter():
    for filename in os.listdir('.'):
        if not filename.endswith('.xlsx'):
            continue
        print('Converting ' + filename + ' into CSV files')
        wb = openpyxl.load_workbook(filename)

        for sheet_names in wb.get_sheet_names():
            sheet = wb.get_sheet_by_name(sheet_names)
            # Get the backward slices for the filename so every file can have a unique  name.
            csv_file = open(filename[: -5] + '_' + sheet_names + '.csv', 'w', newline='')
            csv_writer = csv.writer(csv_file)

            for row_num in range(1, sheet.max_row + 1):
                row_data = []
                for col in range(1, sheet.max_column + 1):
                    data = sheet.cell(row=row_num, column=col).value
                    row_data.append(data)
                for row in row_data:
                    csv_writer.writerow(row)


excel_2_csv_converter()
