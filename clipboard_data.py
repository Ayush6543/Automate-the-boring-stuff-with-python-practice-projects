# ! python3
# clipboard_data.py - Read data from clipboard

import openpyxl, pyperclip

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

clipboard_data = pyperclip.paste()
print(clipboard_data)
for row in range(1, sheet.max_row + 1):
        sheet.cell(row=row, column=3).value = clipboard_data

#wb.save('clipboardData.xlsx')

