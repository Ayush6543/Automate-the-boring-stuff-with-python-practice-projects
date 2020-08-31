# ! python3

"""
spread_sheet_cell_inverter.py - A program to invert values contained in cells.
"""

import openpyxl

wb = openpyxl.load_workbook('spread_sheet_new.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=col, column=row).value = sheet.cell(row=row, column=col).value

wb.save('spread_sheet_.xlsx')
